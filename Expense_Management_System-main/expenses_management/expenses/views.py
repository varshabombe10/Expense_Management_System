from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from .models import Transaction
from .forms import TransactionForm
from django.http import JsonResponse
from django.db.models import Sum
from datetime import datetime
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
import openpyxl
from openpyxl.styles import Border, Side, NamedStyle
from django.http import HttpResponse
from django.utils.timezone import localtime

# Home Page
def home(request):
    return render(request, 'home.html')

# User Registration
def register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        # Check if passwords match
        if password != confirm_password:
            messages.error(request, "Passwords do not match!")
            return redirect("register")

        # Check if username already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username is already taken!")
            return redirect("register")

        # Check if email already exists
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email is already in use!")
            return redirect("register")

        # Create user and log in
        user = User.objects.create_user(username=username, email=email, password=password)
        login(request, user)
        messages.success(request, "Registration successful! Welcome, " + username)
        return redirect("dashboard")

    return render(request, "register.html")

# User Login
def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
    return render(request, 'login.html')

# User Logout
def user_logout(request):
    logout(request)
    return redirect('home')

# Dashboard
@login_required
def dashboard(request):
    transactions = Transaction.objects.filter(user=request.user)
    total_credits = sum(t.amount for t in transactions if t.transaction_type == 'credit')
    total_debits = sum(t.amount for t in transactions if t.transaction_type == 'debit')
    balance = total_credits - total_debits

    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            return redirect('dashboard')
    else:
        form = TransactionForm()

    return render(request, 'dashboard.html', {
        'balance': balance,
        'transactions': transactions,
        'form': form,
        'total_credits': total_credits,
        'total_debits': total_debits
    })

# API endpoint for Chart.js
@login_required
def transaction_chart_data(request):
    transactions = Transaction.objects.filter(user=request.user)

    # Get selected month from request (default: current month)
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    year, month = map(int, month.split('-'))

    # Filter transactions for the selected month
    monthly_transactions = transactions.filter(date__year=year, date__month=month)

    credits = sum(t.amount for t in monthly_transactions if t.transaction_type == 'credit')
    debits = sum(t.amount for t in monthly_transactions if t.transaction_type == 'debit')

    data = {
        "labels": ["Credits", "Debits"],
        "values": [credits, debits]
    }
    return JsonResponse(data)

# Edit Transaction
@login_required
def edit_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)

    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = TransactionForm(instance=transaction)

    return render(request, 'edit_transaction.html', {'form': form})

# Delete Transaction
@login_required
def delete_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
    if request.method == 'POST':
        transaction.delete()
        return redirect('dashboard')

    return render(request, 'delete_transaction.html', {'transaction': transaction})

# Profile
@login_required
def profile(request):
    if request.method == "POST":
        user = request.user
        username = request.POST["username"]
        email = request.POST["email"]
        first_name = request.POST.get("first_name", "")
        last_name = request.POST.get("last_name", "")

        # Update user details
        user.username = username
        user.email = email
        user.first_name = first_name
        user.last_name = last_name
        user.save()
        
        messages.success(request, "Profile updated successfully.")

        # Handle password change
        current_password = request.POST.get("current_password")
        new_password = request.POST.get("new_password")
        confirm_new_password = request.POST.get("confirm_new_password")

        if current_password and new_password:
            if new_password == confirm_new_password:
                if user.check_password(current_password):
                    user.set_password(new_password)
                    user.save()
                    update_session_auth_hash(request, user)  # Keep user logged in
                    messages.success(request, "Password changed successfully.")
                else:
                    messages.error(request, "Current password is incorrect.")
            else:
                messages.error(request, "New passwords do not match.")

        return redirect("dashboard")

    return render(request, "profile.html")

# Excel Download
# Define a date style for the Excel file
date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD HH:MM:SS')

# Define border style
border_style = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

@login_required
def export_transactions(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    headers = ['ID', 'Date', 'Description', 'Amount', 'Tran_Type', 'Payment Method']
    ws.append(headers)

    transactions = Transaction.objects.filter(user=request.user)

    # Add transaction data to the sheet
    for transaction in transactions:
        transaction_date = localtime(transaction.date).replace(tzinfo=None)
        row = [
            transaction.id,
            transaction_date,
            transaction.description,
            transaction.amount,
            transaction.transaction_type,
            transaction.payment_method,  # Added payment method field
        ]
        ws.append(row)

    # Apply the date style to the date column (column 'B')
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # Start from row 2 (skip header)
        for cell in row:
            cell.style = date_style

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border_style

    # Adjust the column width for date column and others
    ws.column_dimensions['B'].width = 20  # Adjust width of the Date column

    # Set the response content type and return the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=transactions.xlsx'

    wb.save(response)
    return response